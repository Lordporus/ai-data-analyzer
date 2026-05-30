from __future__ import annotations

import math
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_excel_report(
    cleaned_df: pd.DataFrame,
    insights: Any,
    forecasts: Any,
    quality_result: Any,
    brand_config: dict | None,
    output_path: str | Path,
) -> str:
    """Create a formatted Excel workbook for cleaned data and analysis outputs."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    brand_color = _normalize_hex((brand_config or {}).get("primary_color", "#6C63FF"))
    header_fill = PatternFill("solid", fgColor=brand_color)
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color=brand_color, bold=True, size=14)
    thin_border = Border(bottom=Side(style="thin", color="D9E2EC"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"
    _write_dataframe(ws, cleaned_df, header_fill, header_font, thin_border)

    kpi_ws = wb.create_sheet("KPI Summary")
    _write_title(kpi_ws, "KPI Summary", title_font)
    kpi_rows = [["Metric", "Value", "Unit", "Description"]]
    for kpi in getattr(insights, "kpi_list", []) or []:
        kpi_rows.append([
            getattr(kpi, "name", ""),
            _safe_excel_value(getattr(kpi, "value", "")),
            getattr(kpi, "unit", ""),
            getattr(kpi, "description", ""),
        ])
    if quality_result is not None:
        kpi_rows.append([
            "Data Quality Score",
            _safe_excel_value(getattr(quality_result, "quality_score", "")),
            "/100",
            "Post-cleaning quality score",
        ])
    _write_rows(kpi_ws, kpi_rows, header_fill, header_font, thin_border, start_row=3)

    insights_ws = wb.create_sheet("Insights")
    _write_title(insights_ws, "Insights And Recommendations", title_font)
    insight_rows = [["Type", "Title", "Evidence", "Recommendation"]]
    for trend in getattr(insights, "trend_summary", []) or []:
        insight_rows.append([
            "Trend",
            getattr(trend, "column", ""),
            f"{getattr(trend, 'direction', '')}; R2={getattr(trend, 'r_squared', '')}",
            getattr(trend, "context", "") or "Review movement over time and compare against business events.",
        ])
    for anomaly in getattr(insights, "anomaly_flags", []) or []:
        insight_rows.append([
            "Anomaly",
            getattr(anomaly, "column", ""),
            f"{getattr(anomaly, 'count', 0)} records flagged",
            "Review flagged records for data entry issues or unusual business events.",
        ])
    for rel in getattr(insights, "key_relationships", []) or []:
        insight_rows.append(["Relationship", str(rel), "", "Use this relationship as a follow-up analysis path."])
    for rec in getattr(insights, "business_recommendations", []) or []:
        insight_rows.append(["Recommendation", "", "", str(rec)])
    _write_rows(insights_ws, insight_rows, header_fill, header_font, thin_border, start_row=3)

    forecast_rows = _forecast_rows(forecasts)
    if forecast_rows:
        forecast_ws = wb.create_sheet("Forecasts")
        _write_title(forecast_ws, "Forecasts", title_font)
        _write_rows(
            forecast_ws,
            [["Metric", "Date", "Forecast", "Lower Bound", "Upper Bound", "Confidence"]] + forecast_rows,
            header_fill,
            header_font,
            thin_border,
            start_row=3,
        )

    for worksheet in wb.worksheets:
        _autosize_columns(worksheet)
        worksheet.freeze_panes = "A2" if worksheet.title == "Cleaned Data" else "A4"

    wb.save(output_path)
    return str(output_path)


def _write_dataframe(ws, df: pd.DataFrame, header_fill, header_font, thin_border) -> None:
    rows = [list(df.columns)]
    rows.extend(df.itertuples(index=False, name=None))
    _write_rows(ws, rows, header_fill, header_font, thin_border)
    if df.columns.any() and len(df) > 0:
        ws.auto_filter.ref = ws.dimensions


def _write_title(ws, title: str, title_font: Font) -> None:
    ws["A1"] = title
    ws["A1"].font = title_font


def _write_rows(ws, rows: Iterable[Iterable[Any]], header_fill, header_font, thin_border, start_row: int = 1) -> None:
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_safe_excel_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r_idx == start_row:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = thin_border


def _forecast_rows(forecasts: Any) -> list[list[Any]]:
    if not forecasts or not getattr(forecasts, "forecasts", None):
        return []

    rows: list[list[Any]] = []
    for metric, data in forecasts.forecasts.items():
        dates = data.get("dates_forecast", []) or data.get("dates", [])
        values = data.get("values_forecast", []) or data.get("values", [])
        lowers = data.get("lower_bound", []) or data.get("lower", [])
        uppers = data.get("upper_bound", []) or data.get("upper", [])
        confidence = data.get("confidence", "")
        for idx, value in enumerate(values):
            rows.append([
                metric,
                dates[idx] if idx < len(dates) else "",
                value,
                lowers[idx] if idx < len(lowers) else "",
                uppers[idx] if idx < len(uppers) else "",
                confidence,
            ])
    return rows


def _safe_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ""
    if isinstance(value, (list, dict, tuple, set)):
        return str(value)
    if isinstance(value, str) and len(value) > 32000:
        return value[:31997] + "..."
    return value


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 10
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 2, 48)


def _normalize_hex(value: str) -> str:
    value = str(value or "#6C63FF").strip().lstrip("#")
    if len(value) != 6:
        return "6C63FF"
    return value.upper()
